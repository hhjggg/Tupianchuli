# -*- coding: utf-8 -*-
"""数据表解析与匹配模块

1. read_order_numbers() : 读取单号表（新建 XLSX 工作表）的“订单编号”列
2. parse_direct_table() : 流式解析直供订单照片导出表（sheet1.xml 约 74MB，
                          openpyxl 太慢，改用 zipfile + ElementTree 流式解析），
                          按“三方单号(16位)==订单编号”匹配，收集照片 URL
3. mark_uploaded()      : 在单号表“是否上传”列、订单编号对应的行写入“是”
"""
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter

from openpyxl import load_workbook

NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
NO_PHOTO_MARK = "暂无国补照片"

# 直供表列位（按单元格引用字母定位）
COL_ORDER_NO = "C"   # 三方单号
COL_PHOTO = "G"      # 国补照片链接
HEADERS = ["订单号", "下单时间", "三方单号", "是否国补订单", "政策地", "承运商", "国补照片链接"]
ALL_COLS = list("ABCDEFG")


def _cell_value(c):
    """读取单元格文本：支持 inlineStr 与普通数值 <v>。"""
    if c.get("t") == "inlineStr":
        is_ = c.find(NS + "is")
        t = is_.find(NS + "t") if is_ is not None else None
        return t.text if t is not None else ""
    v = c.find(NS + "v")
    return v.text if v is not None else None


def _resolve_sheet_xml(zf):
    """定位实际的 worksheet xml 路径（优先 sheet1.xml，兜底用 rels 解析）。"""
    names = zf.namelist()
    candidates = [n for n in names if n.startswith("xl/worksheets/") and n.endswith(".xml")]
    if not candidates:
        raise ValueError("未找到工作表文件")
    if "xl/worksheets/sheet1.xml" in candidates:
        return "xl/worksheets/sheet1.xml"
    try:
        wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
        rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        m = re.search(r'<sheet[^>]*r:id="(rId\d+)"', wb_xml)
        if m:
            rid = m.group(1)
            rm = re.search(r'Id="' + rid + r'"[^>]*Target="([^"]+)"', rels)
            if rm:
                target = rm.group(1)
                if target.startswith("/"):
                    target = target.lstrip("/")
                elif not target.startswith("xl/"):
                    target = "xl/" + target
                if target in names:
                    return target
    except Exception:
        pass
    return candidates[0]


def read_order_numbers(xlsx_path):
    """读取单号表的全部“订单编号”，按原表顺序返回。"""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    header_row, order_col = None, None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if "订单编号" in vals:
            header_row, order_col = i, vals.index("订单编号")
            break
        if i >= 10:
            break
    if header_row is None:
        header_row, order_col = 1, 1  # 兜底：第 1 行表头、第 2 列
    nums = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if order_col < len(row) and row[order_col] is not None:
            s = str(row[order_col]).strip()
            if s:
                nums.append(s)
    wb.close()
    return nums


def parse_direct_table(xlsx_path, order_set):
    """流式解析直供订单照片导出表，返回：
        photo_orders : {订单编号: [照片URL, ...]}  有照片的匹配订单
        seen_orders  : 匹配到的订单编号集合（含无照片）
        matched_rows : 匹配行数
        total_rows   : 数据总行数（不含表头）
        len_counter  : 三方单号长度分布
        filtered_rows: 符合条件的行（三方单号=16位 且 有照片），7列值列表
    """
    zf = zipfile.ZipFile(xlsx_path)
    f = zf.open(_resolve_sheet_xml(zf))
    photo_orders = {}
    seen = set()
    matched_rows = 0
    total_rows = 0
    len_counter = Counter()
    filtered_rows = []
    try:
        for _ev, el in ET.iterparse(f, events=("end",)):
            if el.tag == NS + "row":
                total_rows += 1
                vals = {}
                for c in el.findall(NS + "c"):
                    ref = c.get("r", "")
                    col = "".join(ch for ch in ref if ch.isalpha())
                    vals[col] = _cell_value(c)
                ts = vals.get(COL_ORDER_NO)
                if ts is None:
                    el.clear()
                    continue
                ts = str(ts).strip()
                if ts:
                    len_counter[len(ts)] += 1
                if len(ts) != 16:
                    el.clear()
                    continue
                link = vals.get(COL_PHOTO)
                s = str(link).strip() if link is not None else ""
                has_photo = bool(s) and s != NO_PHOTO_MARK
                if has_photo:
                    filtered_rows.append([vals.get(c, "") if vals.get(c) is not None else "" for c in ALL_COLS])
                if ts in order_set:
                    matched_rows += 1
                    seen.add(ts)
                    if has_photo:
                        photo_orders.setdefault(ts, [])
                        if s not in photo_orders[ts]:
                            photo_orders[ts].append(s)
                el.clear()
    finally:
        f.close()
        zf.close()
    return {
        "photo_orders": photo_orders,
        "seen_orders": seen,
        "matched_rows": matched_rows,
        "total_rows": total_rows,
        "len_counter": len_counter,
        "filtered_rows": filtered_rows,
    }


def merge_direct_tables(file_list, order_set, output_path):
    """处理并拼接多个直供订单表，生成新文件。

    规则：只保留“三方单号=16位 且 有照片”的行；多个文件逐行合并；
    生成的新 xlsx 写入 output_path，同时汇总匹配信息供后续匹配使用。
    返回 dict：photo_orders / seen_orders / matched_rows / total_rows /
               len_counter / kept_rows / per_file / output_path
    """
    from openpyxl import Workbook
    photo_orders = {}
    seen = set()
    matched_rows = 0
    total_rows = 0
    len_counter = Counter()
    all_rows = []
    per_file = []
    for f in file_list:
        res = parse_direct_table(f, order_set)
        per_file.append({
            "name": os.path.basename(str(f)),
            "total": res["total_rows"],
            "kept": len(res["filtered_rows"]),
        })
        total_rows += res["total_rows"]
        matched_rows += res["matched_rows"]
        seen |= res["seen_orders"]
        len_counter += res["len_counter"]
        for ts, urls in res["photo_orders"].items():
            photo_orders.setdefault(ts, [])
            for u in urls:
                if u not in photo_orders[ts]:
                    photo_orders[ts].append(u)
        all_rows.extend(res["filtered_rows"])
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("sheet1")
    ws.append(HEADERS)
    for r in all_rows:
        ws.append(r)
    wb.save(output_path)
    return {
        "photo_orders": photo_orders,
        "seen_orders": seen,
        "matched_rows": matched_rows,
        "total_rows": total_rows,
        "len_counter": len_counter,
        "kept_rows": len(all_rows),
        "per_file": per_file,
        "output_path": output_path,
    }


def mark_uploaded(xlsx_path, order_no):
    """在单号表中把“订单编号 == order_no”行的“是否上传”列写入“是”。

    返回写入的行数；找不到订单编号时返回 0（不保存）。
    """
    wb = load_workbook(xlsx_path)
    try:
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        header_row, up_col, ord_col = None, None, None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if "订单编号" in vals and "是否上传" in vals:
                header_row = i
                up_col = vals.index("是否上传")
                ord_col = vals.index("订单编号")
                break
            if i >= 10:
                break
        if header_row is None or up_col is None or ord_col is None:
            raise ValueError("单号表中未找到“订单编号”或“是否上传”列")
        target = str(order_no).strip()
        count = 0
        for r in range(header_row + 1, ws.max_row + 1):
            cv = ws.cell(row=r, column=ord_col + 1).value
            if cv is None:
                continue
            if str(cv).strip() == target:
                ws.cell(row=r, column=up_col + 1).value = "是"
                count += 1
        if count:
            wb.save(xlsx_path)
        return count
    finally:
        wb.close()


if __name__ == "__main__":
    import time

    here = os.path.dirname(os.path.abspath(__file__))
    f1 = os.path.join(here, "新建 XLSX 工作表.xlsx")
    f2 = os.path.join(here, "直供订单照片导出_1538228250_2026年08月12日18时47分48秒.xlsx")

    t0 = time.time()
    orders = read_order_numbers(f1)
    print(f"[{time.time()-t0:.1f}s] 单号表订单编号 {len(orders)} 个")
    t0 = time.time()
    res = parse_direct_table(f2, set(orders))
    print(f"[{time.time()-t0:.1f}s] 总行数={res['total_rows']} 匹配行={res['matched_rows']} "
          f"有照片订单={len(res['photo_orders'])} "
          f"无照片订单={len(res['seen_orders']-set(res['photo_orders']))}")

